#########################################
# Author: Likun@stu.zzu.edu.cn
# Date: 2015-10-14
# 
# read excel file and transform to txt file
#########################################

from openpyxl import Workbook
from openpyxl import load_workbook

wb = load_workbook(filename = r'/home/lenovo/code/python/LSM/data/output.xlsx')

sheetnames = wb.get_sheet_names()  
ws = wb.get_sheet_by_name(sheetnames[0])  

file_train_one_path = '/home/lenovo/code/python/LSM/data/trainOutOne.txt'
file_train_one  =  open(file_train_one_path,'w')

file_train_two_path = '/home/lenovo/code/python/LSM/data/trainOutTwo.txt'
file_train_two  =  open(file_train_two_path,'w')

file_test_one_path = '/home/lenovo/code/python/LSM/data/testOutOne.txt'
file_test_one  =  open(file_test_one_path,'w')

file_test_two_path = '/home/lenovo/code/python/LSM/data/testOutTwo.txt'
file_test_two  =  open(file_test_two_path,'w')



for row in ws.iter_rows('A2:A4166'):
    for cell in row:
        print cell.value
        file_train_one.write(str(cell.value))
    file_train_one.write('\n')
file_train_one.close()

for row in ws.iter_rows('A4167:A5876'):
    for cell in row:
        print cell.value
        file_test_one.write(str(cell.value))
    file_test_one.write('\n')
file_test_one.close()

for row in ws.iter_rows('B2:B4166'):
    for cell in row:
        print cell.value
        file_train_two.write(str(cell.value))
    file_train_two.write('\n')
file_train_two.close()

for row in ws.iter_rows('B4167:B5876'):
    for cell in row:
        print cell.value
        file_test_two.write(str(cell.value))
    file_test_two.write('\n')
file_test_two.close()
